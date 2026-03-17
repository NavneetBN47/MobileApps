import os
import json
import codecs
import openpyxl
import plistlib
from bs4 import BeautifulSoup
from datetime import datetime
import xml.etree.ElementTree as ET
from SAF.misc import package_utils
from SAF.misc.couch_wrapper import CouchWrapper

from io import open
try:
    import html
    from html.parser import HTMLParser
except ImportError:
    from html.parser import HTMLParser
    html = HTMLParser()

class MyHTMLParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        self.recording = 0
        self.data = []

    def handle_data(self, data):
        self.data.append(data)

class StringProcessor(object):

    lang_list = ["bg", "cs", "da", "de", "el", "en", "es", "et", "fi", "fr", "hr", "hu", "it", "ja", "ko", "lt", "lv", "nb", "nl", "pl", "ro", "ru", "sk", "sl", "sv", "tr"]
    android_ch = ["pt", "zh-rCN", "zh-rTW"]
    ios_ch = ["pt-BR", "zh-Hans", "zh-Hant"]
    windows_lang = ["ar-SA", "he-IL", "pt-BR", "pt-PT", "th-TH", "uk-UA", "zh-HK", "zh-Hans-CN", "zh-Hant-TW"]
    pkg_app_string_dict = {}
    mfe_spec_string_dict = {}
    spec_string_dict = {}


    def __init__(self, os_type, db_connection):
        self.os_type = os_type
        self.db_connection = db_connection
        if self.os_type == "android":
            self.lang_list += self.android_ch
        elif self.os_type == "ios":
            self.lang_list += self.ios_ch
        elif self.os_type == "windows":
            self.lang_list += self.windows_lang


    def get_new_rec_id(self, project_name):
        result = self.db_connection.getViewResultByKey("get_info", "max_id", startKey=[self.os_type, project_name], endKey=[self.os_type, project_name], reduce=True)
        if len(result.rows) == 1:
            return int(result.rows[0]["value"]) + 1
        elif len(result.rows) == 0:
            return 0

    def pull_previous_records(self):
        pass

    def upload_record_to_couchdb(self, project_name, spec_version, pkg_source, apk_app_string_dict, spec_string_dict, result_dict, str_id_total_diff=None):
        rec= {
            "spec_version": spec_version,
            "project_name": project_name,
            "pkg_source": pkg_source,
            "os_type": self.os_type,
            "reviewed": False,
            "upload_time": datetime.now().strftime("%m/%d/%Y, %H:%M:%S"),
            "rec_id": str(self.get_new_rec_id(project_name)),
            "apk_str_dict": apk_app_string_dict, 
            "spec_str_dict": spec_string_dict, 
            "result_str_dict": result_dict,
            "str_id_total_diff": str_id_total_diff
        }
        self.db_connection.saveRecord(rec)

    def extract_and_build_dicts_from_pkg_and_mfe(self, pkg_location, spec_location, mfe_spec_location):
        '''
        Extraction of apk/ipa 
        Obtaining the package/drop strings & MFEs from apk/ipa
        Building dictionary from both xml/xliff and MFEs
        '''
        if self.os_type=="android":
            self.pkg_app_string_dict, self.extracted_path = self.rip_all_str_id_from_pkg_android(pkg_location)
            self.spec_string_dict = self.android_build_dict_from_xml(spec_location)

            if mfe_spec_location != '': 
                self.mfe_spec_string_dict = self.build_dict_from_spec_mfe(mfe_spec_location)
                self.spec_string_dict.update(self.mfe_spec_string_dict)
            self.mfe_path_in_pkg = os.path.join(str(self.extracted_path),"assets","Weblets","hpx-web-mfe-prod","assets", "mfe", "@hpx-core-experiences")

        elif self.os_type == "ios":
            self.pkg_app_string_dict, self.extracted_path = self.rip_all_str_id_from_pkg_ios(pkg_location)
            self.spec_string_dict = self.ios_build_dict_from_xliff(spec_location)

            if mfe_spec_location != '': 
                self.mfe_spec_string_dict = self.build_dict_from_spec_mfe(mfe_spec_location)
                self.spec_string_dict.update(self.mfe_spec_string_dict)
            self.mfe_path_in_pkg = os.path.join(str(self.extracted_path),"Weblets","webapps","assets","mfe","@hpx-core-experiences")
         
        elif self.os_type == "windows":
            self.pkg_app_string_dict, self.extracted_path = self.rip_all_str_id_from_pkg_windows(pkg_location)
            self.spec_string_dict = self.windows_build_dict_from_xml(spec_location)

            if mfe_spec_location != '':
                self.mfe_spec_string_dict = self.build_dict_from_spec_mfe(mfe_spec_location)
                self.spec_string_dict.update(self.mfe_spec_string_dict)
            self.mfe_path_in_pkg = os.path.join(str(self.extracted_path), "Weblets", "webapps", "root-web-mfe", "assets", "mfe", "%40hpx-core-experiences")

        elif self.os_type == "other":
            self.pkg_app_string_dict = self.medillia_build_dict_form_json_source(pkg_location)
            self.spec_string_dict = self.medillia_build_dict_from_xlsx(spec_location)

        return self.mfe_path_in_pkg


    def update_mfe_to_dict(self, selected_mfe, mfe_path_in_pkg):
        '''
        Selected MFEs from GUI is updated to the dictionary for testing
        '''
        self.pkg_app_string_dict.update(self.get_mfe_from_pkg(selected_mfe, mfe_path_in_pkg))
        return self.pkg_app_string_dict


    def run_results_and_upload_to_db(self, project_name, pkg_location, spec_location, test_run=True, include_previous=True):
        '''
        Package(apk/ipa) and spec(XML and/or MFE) dictionaries are compared and the resulting dict is uploaded to couch db if not test_run. 
        If test_run is true, the resulting dict is returned for testing purposes without uploading to couch db
        '''
        results_dict = self.run_results(project_name, self.pkg_app_string_dict, self.spec_string_dict, include_previous=include_previous)
        if include_previous:
            prev_record = self.load_prev_results(project_name)
            prev_apk_dict = prev_record.get("apk_str_dict", {}) if isinstance(prev_record, dict) else {} # Defensive: handle case where prev_record is None or not a dict
            if isinstance(prev_apk_dict, dict):
                str_id_diff_total = len(self.pkg_app_string_dict.keys()) - len(prev_apk_dict.keys())
            else:
                str_id_diff_total = None
        else:
            str_id_diff_total = None

        if not test_run:
            spec_version = spec_location.split("/")[-1]
            self.upload_record_to_couchdb(project_name, spec_version, pkg_location, self.pkg_app_string_dict, self.spec_string_dict, results_dict, str_id_total_diff=str_id_diff_total)
        else:
            # Return both keys for compatibility with old/new GUI code
            return {
                "result_dict": results_dict,
                "result_str_dict": results_dict,
                "spec_total": len(self.spec_string_dict.keys()),
                "str_id_total_diff": str_id_diff_total
            }


    def medillia_build_dict_from_xlsx(self, xlsx_path):
        data = {}
        book = openpyxl.load_workbook(xlsx_path)
        for item in book.sheetnames:
            sheet = book[item]
            id = 0
            language = None
            for column in range(sheet.max_column):
                id = 0
                for row in range(sheet.max_row): 
                    if row==0:
                        language = sheet.cell(row=row+1, column=column+1).value.split(" ")[-1].lower()
                        continue
                    elif row in [1, 2]:
                        continue
                    str_id = item+"_"+str(id)
                    if sheet.cell(row=row+1, column=column+1).value is None:
                        break
                    if data.get(str_id, None) is None:
                        data[str_id] = {language:sheet.cell(row=row+1, column=column+1).value.strip()}
                    else:
                        data[str_id][language] = sheet.cell(row=row+1, column=column+1).value.strip()
                    id += 1
        return data

    def medillia_build_dict_form_json_source(self, json_path):
        parser = MyHTMLParser()
        fh = open(json_path, "r", encoding="utf-8")
        data_file = json.load(fh)
        return_data = {}
        for _a in data_file["propertyConfiguration"]["forms"]:
            gi = False
            if "TEST" in _a["formJson"]["name"]:
                continue
            if "General Intercept" in _a["formJson"]["name"]:
                gi = True
                sheet_name = "General Intercept"
            else: 
                sheet_name = _a["formJson"]["name"].split(" ")[-2]
            id = 0
            language = _a["formJson"]["name"].split(" ")[-1].lower()
            got_placeholder = False
    
            actual_id = sheet_name+"_"+str(id)

            for _b in _a["formJson"]["pages"]:

                for _c in _b["dynamicData"]:
                    if _c["labelContent"] == "":
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["label"].strip()
                        id += 1
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)
                    else:
                        parser.data=[]
                        parser.feed(html.unescape(_c["labelContent"]))
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: "".join(parser.data).strip("\n").strip()}
                        else:
                            return_data[actual_id][language] = "".join(parser.data).strip("\n").strip()
                        id += 1   
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                    if _c["optionsById"] is not None:
                        for d in _c["optionsById"]:
                            if return_data.get(actual_id, None) is None:
                                return_data[actual_id] = {language: d["label"].strip()}
                            else:
                                return_data[actual_id][language] = d["label"].strip()
                            id += 1 
                            if id == 28:
                                id+=1
                            actual_id = sheet_name+"_"+str(id)
                    if _c.get("ratingScales", None) is not None:
                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["ratingScales"][0]["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["ratingScales"][0]["label"].strip()
                        id += 1 
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                        if return_data.get(actual_id, None) is None:
                            return_data[actual_id] = {language: _c["ratingScales"][-1]["label"].strip()}
                        else:
                            return_data[actual_id][language] = _c["ratingScales"][-1]["label"].strip()
                        id += 1 
                        if id == 28:
                            id+=1
                        actual_id = sheet_name+"_"+str(id)

                    if not got_placeholder and _c.get("placeholder", "") not in ["", "Insert label text"]:

                        if return_data.get(actual_id, None) is None:
                            return_data[sheet_name+"_"+str(28)] = {language: _c["placeholder"].strip()}
                        else:
                            return_data[sheet_name+"_"+str(28)][language] = _c["placeholder"].strip()
                        got_placeholder = True

            id = 0
            if not gi and _a["formJson"]["settings"].get("formBasicSettings", {}).get("submitButtonLabel", "") != "":
                if return_data.get(actual_id, None) is None:
                    return_data[actual_id] = {language: _a["formJson"]["settings"]["formBasicSettings"]["submitButtonLabel"]}
                else:
                    return_data[actual_id][language] = _a["formJson"]["settings"]["formBasicSettings"]["submitButtonLabel"]
                id += 1 
                if id == 28:
                    id+=1
                actual_id = sheet_name+"_"+str(id)
            id = 0
            if gi:
                sheet_name += " Invite"
                actual_id = sheet_name+"_"+str(id)
                invite_field_keys = ["invitationHeadline","invitationText","provideButtonText","declineButtonText","laterButtonText"]
                for key in invite_field_keys:
                    if return_data.get(actual_id, None) is None:
                        return_data[actual_id] = {language: _a["inviteData"][key].strip()}
                    else:
                        return_data[actual_id][language] = _a["inviteData"][key].strip()
                    id += 1 
                    if id == 28:
                        id+=1
                    actual_id = sheet_name+"_"+str(id)
        return return_data

    def rip_all_str_id_from_pkg_windows(self, pkg_location):
        '''
        Extraction of Windows package and obtaining dictionary containing the app strings
        '''
        base_dir = pkg_location
        if not os.path.isdir(base_dir):
            if os.path.isfile(base_dir):
                base_dir = os.path.dirname(base_dir)
            else:
                raise OSError("Windows pkg_location must be a directory or bundle file path: " + str(pkg_location))

        extracted_path = os.path.join(base_dir, "app_extracted")
        method_dict = {"windows": package_utils.get_app_string_from_windows_bundle}

        pkg_app_string_dict = {}
        for lang in self.lang_list:
            localized_str = method_dict[self.os_type](pkg_location, lang)
            language = lang.split("/")[-1]
            for key, value in localized_str.items():
                if pkg_app_string_dict.get(key, None) is None:
                    pkg_app_string_dict[key] = {language:value}
                else:
                    pkg_app_string_dict[key][language] = value

        # Try to point extracted_path at the main extracted package folder
        for arch in ("x64", "x86", "ARM"):
            candidate = os.path.join(extracted_path, arch)
            if os.path.isdir(candidate):
                extracted_path = candidate
                print("App package extracted path is: " + extracted_path)
                break
        return pkg_app_string_dict, extracted_path
    
    def windows_build_dict_from_xml(self, app_strings_folder):
        """Building dictionary from XML files in the spec location for Windows. 
        The XML files are expected to be in the format of {lang}.xml (e.g. en.xml, de.xml etc.)
        paramm app_strings_folder: the folder path containing the XML files for Windows spec strings
        return: a dictionary in the format of {str_id: {lang: localized_string}} containing the localized strings from the XML files in spec location
        """
        if not os.path.isdir(app_strings_folder):
            raise OSError("Path: " + str(app_strings_folder) + " is not a directory")

        spec_app_string_dict = {}
        for lang in self.lang_list:  
            localized_str = package_utils.windows_read_app_string_from_xml(None, app_strings_folder, lang)
            language = lang.split("/")[-1]
            for key, value in localized_str.items():
                if spec_app_string_dict.get(key, None) is None:
                    spec_app_string_dict[key] = {language: value}
                else:
                    spec_app_string_dict[key][language] = value
        return spec_app_string_dict

    def rip_all_str_id_from_pkg_android(self, pkg_location):
        '''
        Extraction of APK and and obtaining dictionary containing the app strings
        '''
        _, package_name = package_utils.check_if_pkg_exist(pkg_location)
        extracted_path = "/tmp/temp_app_extracted/" + package_name + "/"
        method_dict = {"android": package_utils.get_app_string_from_apk}

        pkg_app_string_dict = {}
        for lang in self.lang_list:
            localized_str = method_dict[self.os_type](pkg_location, lang, eng_base=False)
            language = lang.split("/")[-1]
            for key, value in localized_str.items():
                if pkg_app_string_dict.get(key, None) is None:
                    pkg_app_string_dict[key] = {language:value}
                else:
                    pkg_app_string_dict[key][language] = value
        return pkg_app_string_dict, extracted_path

    def android_build_dict_from_xml(self, path):
        spec_app_string_dict = {}
        res_path = path + "/res"
        if not os.path.isdir(path):
            raise OSError("Path: " + path + " is not a directory")
        if not os.path.isdir(res_path):
            raise OSError("Path: " + path + " does not contain the res folder")

        lang_dir = [os.path.join(res_path,x) for x in os.listdir(res_path) if os.path.isdir(os.path.join(res_path, x))]
        for lang in lang_dir:
            language = "-".join(lang.split("/")[-1].split("-")[1:]) if "-" in lang else "en"
            all_files = [os.path.join(lang,x) for x in os.listdir(lang)]
            lang_files = []
            for files in all_files:
                files_parts = files.split("/")[-1].split(".")[0].split("_")
                if files_parts[0] == "strings":
                    lang_files.append(files)
                else:
                    os.remove(files)
            for xml_files in lang_files:
                tree = ET.parse(xml_files)
                root = tree.getroot()
                for str_id in root.findall('string'):
                    if spec_app_string_dict.get(str_id.get("name"), {}).get(language, False):
                        raise ValueError("Error: " + str_id["name"] + " duplicated")
                    raw_string = u"".join([x for x in str_id.itertext()]).rstrip()
                    unescaped_str = codecs.escape_decode(bytes(raw_string, "utf-8"))[0].decode("utf-8").strip()
                    if not spec_app_string_dict.get(str_id.get("name"), False):
                        spec_app_string_dict[str_id.get("name")] = {language: unescaped_str}
                    else:
                        spec_app_string_dict[str_id.get("name")][language] = unescaped_str
        return spec_app_string_dict

    def rip_all_str_id_from_pkg_ios(self, pkg_location):
        '''
        Extraction of IPA and and obtaining dictionary containing the app strings and HTML
        '''
        method_dict = {"ios": package_utils.get_app_string_from_ipa}
        pkg_app_string_dict = {}
        for lang in self.lang_list:
            localized_str = method_dict[self.os_type](pkg_location, lang, eng_base=False)
            language = lang.split("/")[-1]
            for key, value in localized_str.items():
                if pkg_app_string_dict.get(key, None) is None:
                    pkg_app_string_dict[key] = {language:value}
                else:
                    pkg_app_string_dict[key][language] = value

        langs = []
        html_string_dict = {}
        html_files = set()
        bs_text = ""

        _, package_name = package_utils.check_if_pkg_exist(pkg_location)
        extracted_path = "/tmp/temp_app_extracted/" + package_name + "/"

        for lang_files in os.listdir(extracted_path):
            if lang_files.endswith('.lproj'):
                lang_name = lang_files.split(".")[0]
                langs.append(lang_name)

        for x in langs:
            res_path = extracted_path + f"/{x}.lproj/"
            for files in os.listdir(res_path):
                if files.endswith('.strings'):
                    if files.startswith('Localizable'):
                        continue
                    file = os.path.join(res_path, files)
                    with open(file, 'rb') as f:
                        plist_data = plistlib.load(f)
                        for i, j in plist_data.items():
                            if i not in pkg_app_string_dict:
                                pkg_app_string_dict[i] = {x : j}
                            else:
                                pkg_app_string_dict[i][x] = j

        # implemented BeautifulSoup for extracting the .html files from each .lproj folder
        for x in langs:
            res_path = extracted_path + f"/{x}.lproj/"
            for files in os.listdir(res_path):
                if files.endswith('.html'):
                    html_files.add(files)
                    html_names = list(html_files)
                    html_no_ext = [os.path.splitext(j)[0] for j in html_names]
                    file = os.path.join(res_path, files)
                    with open(file, 'r', encoding='utf-8') as f:
                        html_contents = f.read()
                        bs_text = BeautifulSoup(html_contents, features="html.parser")
                        html_text = bs_text.get_text()
                        cleaned_text = html_text.replace("\n", " ")
                        html_string_dict.update({
                            x : cleaned_text.strip()
                            })
                        for i in html_no_ext:
                            pkg_app_string_dict.update({
                            i : html_string_dict
                            })

        return pkg_app_string_dict, extracted_path

    def ios_build_dict_from_xliff(self, resource_folder):
        spec_app_string_dict = {}
        lang_dir = [os.path.join(resource_folder, x) for x in os.listdir(resource_folder) if os.path.isfile(os.path.join(resource_folder, x))]
        for lang_file in lang_dir:
            language = lang_file.split("/")[-1].split(".")[0]
            print("-----------------------Language file-------------{}".format(lang_file))
            reg = ET.parse(lang_file)
            xmlns = reg.getroot().tag.split("}")[0] + "}"
            re_strings = reg.findall(".//" + xmlns + "trans-unit")
            for item in re_strings:
                target = item.find(xmlns + "target")
                str_txt = codecs.escape_decode(bytes(target.text, "utf-8"))[0].decode("utf-8").strip()
                if spec_app_string_dict.get(item.get("id"), None) is None:
                    spec_app_string_dict[item.get("id")] = {language: str_txt}
                else:
                    spec_app_string_dict[item.get("id")][language] = str_txt
        return spec_app_string_dict

    def get_mfe_from_pkg(self, selected_mfe, mfe_path_in_pkg):
        '''
        This implementation is for both android and ios
        The MFEs contain JSON files of all the languages
        A dictionary is returned containing the translation from all languages for each unique key
        '''
        selected_mfe_dict = {}
        mfe_localization_string = "/assets/locale/"

        for mfe in selected_mfe:
            res_path = mfe_path_in_pkg + "/" + mfe + mfe_localization_string
            langs = []
            for lang_files in os.listdir(res_path):
                if lang_files.endswith('.json'):
                    language = lang_files.split("_")[0]
                    langs.append(language)
            for files in os.listdir(res_path):
                if files.endswith('.json'):
                    lang_name = files.split('.')[0]
                    file = os.path.join(res_path, files)
                    with open(file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    for i, j in data.items():
                        if j == "":
                            key = i + "$"
                            if key not in selected_mfe_dict:
                                    selected_mfe_dict[key] = {lang_name : j}
                            else:
                                selected_mfe_dict[key][lang_name] = j

                        if type(j) == dict:
                            for k, l in j.items():
                                key = i + "$" + k
                                if key not in selected_mfe_dict:
                                    selected_mfe_dict[key] = {lang_name : l}
                                else:
                                    selected_mfe_dict[key][lang_name] = l

                        if type(j) == str and j == "":
                            s = dict(j)
                            for k, l in s.items():
                                key = i + "$" + k
                                if key not in selected_mfe_dict:
                                    selected_mfe_dict[key] = {lang_name : l}
                                else:
                                    selected_mfe_dict[key][lang_name] = l
                        
                        if type(j)==str and j != "":
                            key = i + "$"
                            if key not in selected_mfe_dict:
                                    selected_mfe_dict[key] = {lang_name : j}
                            else:
                                selected_mfe_dict[key][lang_name] = j
        return selected_mfe_dict

    def build_dict_from_spec_mfe(self, path):
        '''
        This implementation is for extracting spec MFE strings from json(both ios & android)
        The MFEs contain JSON files of all the languages
        A dictionary is returned containing the translation from all languages for each unique key
        '''
        mfe_spec_app_string_dict = {}
        
        mfe_locales_path = "/assets/locale/"
        for mfe in os.listdir(path):
            res_path = path + "/" + mfe + mfe_locales_path
            langs = []
            for lang_files in os.listdir(res_path):
                if lang_files.endswith('.json'):
                    language = lang_files.split("_")[0]
                    langs.append(language)
            for files in os.listdir(res_path):
                if files.endswith('.json'):
                    lang_name = files.split('.')[0]
                    file = os.path.join(res_path, files)
                    with open(file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    for i, j in data.items():
                        if j == "":
                            key = i + "$"
                            if key not in mfe_spec_app_string_dict:
                                    mfe_spec_app_string_dict[key] = {lang_name : j}
                            else:
                                mfe_spec_app_string_dict[key][lang_name] = j

                        if type(j) == dict:
                            for k, l in j.items():
                                key = i + "$" + k
                                if key not in mfe_spec_app_string_dict:
                                    mfe_spec_app_string_dict[key] = {lang_name : l}
                                else:
                                    mfe_spec_app_string_dict[key][lang_name] = l

                        if type(j) == str and j == "":
                            s = dict(j)
                            for k, l in s.items():
                                key = i + "$" + k
                                if key not in mfe_spec_app_string_dict:
                                    mfe_spec_app_string_dict[key] = {lang_name : l}
                                else:
                                    mfe_spec_app_string_dict[key][lang_name] = l
                        
                        if type(j)==str and j != "":
                            key = i + "$"
                            if key not in mfe_spec_app_string_dict:
                                    mfe_spec_app_string_dict[key] = {lang_name : j}
                            else:
                                mfe_spec_app_string_dict[key][lang_name] = j
        return mfe_spec_app_string_dict

    def load_prev_results(self, project_name):
        prev_rec_id = self.get_new_rec_id(project_name) - 1
        db_rec = self.db_connection.getViewResultByKey("get_info", "rec_by_id", startKey=[self.os_type, project_name, str(prev_rec_id)], endKey=[self.os_type, project_name, str(prev_rec_id)])
        if len(db_rec.rows) != 1:
            #print "Parent record doesn't exist: " + str(prev_rec_id) + " " + self.os_type
            return None
        else:
            #print "Parent record found"
            return db_rec.rows[0]["value"]

    def run_results(self, project_name, apk_app_string_dict, spec_app_string_dict, include_previous=True):
        """Run the results comparison for the given project and app string dictionaries."""
        results_dict = {}
        if include_previous:
            prev_rec = self.load_prev_results(project_name)
            if isinstance(prev_rec, dict): # Defensive: handle case where prev_rec is None or not a dict
                prev_results_dict = prev_rec.get("result_str_dict")
                if not isinstance(prev_results_dict, dict):
                    prev_results_dict = prev_rec.get("result_dict")

                prev_spec_dict = prev_rec.get("spec_str_dict", {})
                if not isinstance(prev_spec_dict, dict):
                    prev_spec_dict = {}

                prev_apk_dict = prev_rec.get("apk_str_dict", {})
                if not isinstance(prev_apk_dict, dict):
                    prev_apk_dict = {}

                if isinstance(prev_results_dict, dict):
                    for str_id, info in prev_results_dict.items():
                        if not isinstance(info, dict):
                            continue

                        info_result = info.get("result")
                        if info_result == "passed" or info_result == "removed" or info_result == "not test":
                            parent_spec_dict = prev_spec_dict.get(str_id, None)
                            parent_pkg_dict = prev_apk_dict.get(str_id, None)
                            cur_spec_dict = spec_app_string_dict.get(str_id, None)
                            cur_pkg_dict = apk_app_string_dict.get(str_id, None)

                            if parent_pkg_dict == cur_pkg_dict and parent_spec_dict == cur_spec_dict:
                                if parent_pkg_dict is None:
                                    results_dict[str_id] = {"result": info_result, "reason": "[3][" + info_result + "]Legacy: str_id not in apk", "reviewed": True}
                                elif parent_spec_dict is None:
                                    results_dict[str_id] = {"result": info_result, "reason": "[4][" + info_result + "]Legacy: str_id not in spec", "reviewed": True}
                                else:
                                    results_dict[str_id] = {"result": info_result, "reason": "[2][" + info_result + "]Legacy: str data unchanged", "reviewed": True}
                        elif info_result == "failed" and info.get("reviewed", False) is True:
                            parent_spec_dict = prev_spec_dict.get(str_id, None)
                            parent_pkg_dict = prev_apk_dict.get(str_id, None)
                            cur_spec_dict = spec_app_string_dict.get(str_id, None)
                            cur_pkg_dict = apk_app_string_dict.get(str_id, None)
                            results_dict[str_id] = {"result": info_result, "reason": str(info.get("reason", "")) + " (legacy)", "reviewed": True}

        for str_id in apk_app_string_dict.keys():
            if include_previous and str_id in results_dict.keys():
                #print "Result migrated from previous run"
                continue
            if not spec_app_string_dict.get(str_id, False):
                if len(apk_app_string_dict[str_id].keys()) > 1:
                    results_dict[str_id] ={"result": "failed", "reason": "[3][failed]Has translation but not in spec"}
                    continue
                results_dict[str_id]={"result": "unknown", "reason": "[1][unknown]Missing in spec"}
                continue

            apk_lang_list = apk_app_string_dict[str_id].keys()
            spec_lang_list = spec_app_string_dict[str_id].keys()

            if set(apk_lang_list) != set(spec_lang_list):
                if len(apk_lang_list) > len(spec_lang_list):
                    #print "Failed: " + str_id + " missing lang in spec: " + str([item for item in apk_lang_list if item not in spec_lang_list])
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in spec: " + str(set(apk_lang_list)-set(spec_lang_list))}
                    continue
                elif len(apk_lang_list) < len(spec_lang_list):
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in apk: " + str(set(spec_lang_list)-set(apk_lang_list))}
                    continue
                else:
                    #If they are the same length but the content are different
                    results_dict[str_id] = {"result": "failed", "reason": "[2][failed]Missing lang in apk: " + str(set(spec_lang_list)-set(apk_lang_list)) + " and Missing lang in spec: " + str(set(apk_lang_list)-set(spec_lang_list))}
                    continue

            if spec_app_string_dict[str_id] != apk_app_string_dict[str_id]:
                mismatch_lang = []
                for lang in apk_app_string_dict[str_id].keys():
                    if apk_app_string_dict[str_id][lang] != spec_app_string_dict[str_id][lang]:
                        mismatch_lang.append(lang)

                results_dict[str_id] = {"result": "failed", "reason": "[1][failed]Mismatch: " + str(mismatch_lang)}
                continue
            else:
                english = apk_app_string_dict[str_id].get("en", None)
                passed_but_empty = []
                passed_but_same = []
                for lang in apk_app_string_dict[str_id].keys():       
                    if apk_app_string_dict[str_id][lang] == "":
                        passed_but_empty.append(lang)
                    elif lang != "en" and english is not None and english == apk_app_string_dict[str_id][lang]:
                        passed_but_same.append(lang)

                if passed_but_empty == apk_app_string_dict[str_id].keys():
                    results_dict[str_id] = {"result": "passed", "reason": "[5][passed]Empty Strings: " + "lang: " + str(passed_but_empty), "reviewed": True}
                elif passed_but_same != []:
                    results_dict[str_id] = {"result": "failed", "reason": "[4][failed]Lang: " + str(passed_but_same) + " is the same as en"}
                else:
                    results_dict[str_id] = {"result": "passed", "reason": "[1][passed]Everything match during current run", "reviewed": True}

        return results_dict
